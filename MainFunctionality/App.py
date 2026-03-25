import tkinter as tk
import sys
from tkinter import ttk, messagebox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import numpy as np
import importlib
import ValueGenerator
from tkinter.filedialog import asksaveasfilename
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import os

# app.py

class InterpolationApp:
    def __init__(self, root):
        self.root = root 
        self.root.title("Interpolation with Variance")
        self.entries = []  # list of (label_widget, entry_widget)
        self.current_values = []  # list of numeric values from backend
        self.current_bounds = []  # list of bounds for editing
        self.graph_title = tk.StringVar()  # graph title
        self.selected_point = None  # index of selected point for editing
        self.child_windows = []
        self.root.protocol("WM_DELETE_WINDOW", self.on_app_close)

        # State for the main graph window
        self.main_graph_state = {
            "parent": self.root,
            "ax": None,
            "canvas": None,
            "bounds": [],
            "plot_points": [],
            "title": ""
        }

        self._build_ui()
        self._init_plot()

    def _build_ui(self):
        frm = ttk.Frame(self.root, padding=8)
        frm.pack(fill="both", expand=True)

        # Inputs frame
        self.inputs_frame = ttk.LabelFrame(frm, text="Target Values", padding=8)
        self.inputs_frame.pack(side="top", fill="x", expand=False)

        # Fields
        self._add_entry("Start")
        self._add_entry("End")
        self._add_entry("Steps")
        self._add_entry("Graph Title")

        # Controls
        controls = ttk.Frame(frm)
        controls.pack(side="top", fill="x", pady=(8, 0))

        submit_btn = ttk.Button(controls, text="Submit", command=self.submit)
        submit_btn.pack(side="left", padx=(8, 0))

        clear_btn = ttk.Button(controls, text="Clear", command=self.clear_all)
        clear_btn.pack(side="left", padx=(8, 0))

        copy_btn = ttk.Button(controls, text="Copy Values", command=lambda: self.copy_to_clipboard(self.main_graph_state))
        copy_btn.pack(side="left", padx=(8, 0))

        new_graph_btn = ttk.Button(controls, text="New Graph", command=self.new_graph)
        new_graph_btn.pack(side="left", padx=(8, 0))

        save_btn = ttk.Button(controls, text="Save Graph", command=lambda: self.save_graph(self.current_values, self.current_bounds, self.graph_title))
        save_btn.pack(side="left", padx=(8, 0))

        # Info label for editing
        self.info_label = ttk.Label(controls, text="Click points on graph to edit", foreground="gray")
        self.info_label.pack(side="left", padx=(20, 0))

    def _add_entry(self, label_text, before_end=False):
        # Insert after the first entry if before_end=True, otherwise append
        if before_end and len(self.entries) >= 1:
            insert_index = 1
        else:
            insert_index = len(self.entries)

        # create widgets
        label = ttk.Label(self.inputs_frame, text=f"{label_text}:")
        entry = ttk.Entry(self.inputs_frame, width=12)

        # grid placement: simple vertical stack
        # First, shift down widgets below insert index
        for i in range(insert_index, len(self.entries)):
            lbl, ent = self.entries[i]
            lbl.grid_configure(row=i + 1, column=0, sticky="w", pady=2)
            ent.grid_configure(row=i + 1, column=1, sticky="w", pady=2)

        label.grid(row=insert_index, column=0, sticky="w", pady=2, padx=(0, 6))
        entry.grid(row=insert_index, column=1, sticky="w", pady=2)

        self.entries.insert(insert_index, (label, entry))

        # Pre-fill placeholder for convenience
        if label_text == "Start":
            entry.insert(0, self.current_bounds[0] if self.current_bounds else "10")
        elif label_text == "End" and entry.get() == "":
            entry.insert(0, self.current_bounds[1] if self.current_bounds else "90")
        elif label_text == "Steps" and entry.get() == "":
            entry.insert(0, self.current_bounds[2] if self.current_bounds else "24")

    def _init_plot(self):
        self.fig = Figure(figsize=(6, 3), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_xlabel("X")
        self.ax.set_ylabel("Value")

        canvas_frame = ttk.LabelFrame(self.root, text="Plot", padding=6)
        canvas_frame.pack(fill="both", expand=True, padx=8, pady=8)

        self.canvas = FigureCanvasTkAgg(self.fig, master=canvas_frame)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack(fill="both", expand=True)

        # Store references in main graph state
        self.main_graph_state["ax"] = self.ax
        self.main_graph_state["canvas"] = self.canvas

        # Connect click event for point editing
        self.canvas.mpl_connect(
            "button_press_event",
            lambda event: self.on_plot_click(event, self.main_graph_state)
        )

    def _clear_plot(self):
        self.ax.cla()
        self.ax.set_xlabel("X")
        self.ax.set_ylabel("Value")
        self.canvas.draw()

    def copy_to_clipboard(self, graph_state):
        if not graph_state["plot_points"]:
            messagebox.showinfo("No values", "There are no values to copy. Please submit first.")
            return
        text = " ".join(str(v) for v in graph_state["plot_points"])
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        
        # # Shows brief confirmation message, could be improved with a temporary label or status bar in the future, or could be entirely removed if deemed unnecessary.
        # messagebox.showinfo("Copied", "Interpolated values copied to clipboard.")

    def new_graph(self):
        """Open a new window to plot and edit a separate graph."""
        self.copy_to_clipboard(self.main_graph_state)  # Copy current values for easy pasting into new graph

        new_window = tk.Toplevel(self.root)
        self.child_windows.append(new_window) # Clean close
        new_window.protocol("WM_DELETE_WINDOW", lambda w=new_window: self.close_child_window(w)) # Clean close
        new_window.title("Multi-Plot Viewer")
        new_window.geometry("700x700")

        # Frame for controls
        control_frame = ttk.Frame(new_window, padding=8)
        control_frame.pack(side="top", fill="x")

        ttk.Label(control_frame, text="Paste clipboard data (one series per line):").pack(anchor="w", pady=(0, 5))

        # Text widget for input
        text_widget = tk.Text(control_frame, height=6, width=50)
        text_widget.pack(fill="both", expand=False, pady=(0, 8))

        # Try to load from clipboard
        try:
            clipboard_data = new_window.clipboard_get()
            text_widget.insert("1.0", clipboard_data)
        except Exception:
            pass

        # Figure and canvas for this new window
        fig = Figure(figsize=(6, 4), dpi=100)
        ax = fig.add_subplot(111)

        canvas_frame = ttk.LabelFrame(new_window, text="Plot", padding=6)
        canvas_frame.pack(fill="both", expand=True, padx=8, pady=8)

        canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True)

        # This graph window's independent state
        graph_state = {
            "parent": new_window,
            "ax": ax,
            "canvas": canvas,
            "bounds": [],
            "plot_points": [],
            "title": self.graph_title.get() or "Multi-Series Plot",
            "text_widget": text_widget
        }

        def plot_data():
            text = text_widget.get("1.0", tk.END).strip()
            if not text:
                messagebox.showwarning("No data", "Please paste data to plot.")
                return

            try:
                values = []
                for line in text.split("\n"):
                    if line.strip():
                        values.extend([int(x.strip()) for x in line.split() if x.strip()])
            except ValueError:
                messagebox.showerror("Invalid format", "Could not parse input.")
                return

            if not values:
                messagebox.showwarning("No data", "No numeric values were found to plot.")
                return

            bounds = [values[0], values[-1], len(values)]

            graph_state["bounds"] = bounds[:]
            graph_state["plot_points"] = values[:]
            graph_state["title"] = self.graph_title.get() or "Multi-Series Plot"

            self.render_plot(
                graph_state["ax"],
                graph_state["canvas"],
                graph_state["bounds"],
                graph_state["plot_points"],
                graph_state["title"]
            )

        # Bind clicks to THIS canvas, for THIS graph only
        canvas.mpl_connect(
            "button_press_event",
            lambda event, state=graph_state: self.on_plot_click(event, state)
        )

        ttk.Button(control_frame, text="Plot Data", command=plot_data).pack(side="left", padx=(0, 5))
        ttk.Button(control_frame, text="Copy Values", command=lambda: self.copy_to_clipboard(graph_state)).pack(side="left", padx=(0, 5))
        ttk.Button(
            control_frame,
            text="Save Graph",
            command=lambda: self.save_graph(
                graph_state["plot_points"],
                graph_state["bounds"],
                tk.StringVar(value=graph_state["title"])
            )
        ).pack(side="left", padx=(0, 5))

        plot_data()  # auto-plot on open if clipboard data is available

    def on_plot_click(self, event, graph_state):
        """Handle clicks on plot points for editing."""
        ax = graph_state["ax"]
        plot_points = graph_state["plot_points"]

        if event.inaxes != ax or not plot_points or event.xdata is None or event.ydata is None:
            return

        clicked_x, clicked_y = event.xdata, event.ydata

        x_interp = np.arange(len(plot_points))
        y_interp = np.array(plot_points)

        distances = np.sqrt((x_interp - clicked_x) ** 2 +
                            (y_interp - clicked_y) ** 2)

        idx = np.argmin(distances)

        if distances[idx] < 1:
            self.edit_value(idx, graph_state)

    def edit_value(self, idx, graph_state):
        """Open dialog to edit a specific value."""
        dialog = tk.Toplevel(graph_state["parent"])
        dialog.title(f"Edit Value {idx + 1}")
        dialog.geometry("300x150")

        label = ttk.Label(dialog, text=f"Edit value for point {idx + 1}:")
        label.pack(pady=10)
        label.configure(font=("Arial", 12, "bold"))

        entry = ttk.Entry(dialog, width=15)
        entry.pack(pady=5)
        entry.insert(0, str(graph_state["plot_points"][idx]))
        entry.focus()

        def save():
            try:
                new_val = int(entry.get().strip())
                graph_state["plot_points"][idx] = new_val
                dialog.destroy()

                # Keep main graph data in sync if editing the main graph
                if graph_state is self.main_graph_state:
                    self.current_values = graph_state["plot_points"]
                    self.current_bounds = graph_state["bounds"]

                # Keep the New Graph text box synced with edited values
                if "text_widget" in graph_state and graph_state["text_widget"] is not None:
                    text_widget = graph_state["text_widget"]
                    text_widget.delete("1.0", tk.END)
                    text_widget.insert("1.0", "\n".join(str(v) for v in graph_state["plot_points"]))

                self.render_plot(
                    graph_state["ax"],
                    graph_state["canvas"],
                    graph_state["bounds"],
                    graph_state["plot_points"],
                    graph_state["title"]
                )
            except ValueError:
                messagebox.showerror("Invalid input", "Please enter a numeric value.")

        ttk.Button(dialog, text="Save", command=save).pack(pady=5)

    

    def submit(self):
        try:
            values = []
            graph_title = None
            for i, (lbl, ent) in enumerate(self.entries):
                txt = ent.get().strip()
                if txt == "":
                    raise ValueError("All fields must be filled.")
                
                # Last entry is Graph Title (text), others are numeric
                if i == len(self.entries) - 1:
                    graph_title = txt
                else:
                    val = int(txt)
                    values.append(val)
        except ValueError as e:
            messagebox.showerror("Invalid input", f"Please enter values in all fields.\n{e}")
            return

        if len(values) < 2:
            messagebox.showerror("Not enough points", "Please provide at least start and end values.")
            return

        # Store graph title for later use
        self.graph_title = tk.StringVar(value=graph_title or "")
        self.plot_results(values, [])

    def save_graph(self, current_values, current_bounds, graph_title):
        """Save the current graph and values to an Excel file"""
        if not current_values:
            messagebox.showinfo("No graph", "There is no graph to save. Please submit first.")
            return
        
        filename = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            # Create workbook and write data
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Interpolation Data"
            
            # Write metadata
            ws['A1'] = "Graph Title"
            ws['B1'] = graph_title.get() if graph_title else "Untitled"
            ws['A2'] = "Start"
            ws['B2'] = current_bounds[0]
            ws['A3'] = "End"
            ws['B3'] = current_bounds[1]
            ws['A4'] = "Steps"
            ws['B4'] = current_bounds[2]
            
            # Write values in columnar format
            ws['A6'] = "Value"
            for i, val in enumerate(current_values):
                ws.cell(row=7+i, column=1).value = val
            
            # Create graph
            fig = Figure(figsize=(10, 6), dpi=100)
            ax = fig.add_subplot(111)
            
            x_out = np.arange(len(current_values))
            y_out = current_values
            ax.plot(x_out, y_out, label="Interpolated (backend)")
            ax.scatter(x_out, y_out, color="red", zorder=5, label="Interpolated points")
            
            start_bound = [0, len(current_values) - 1]
            end_bound = [current_bounds[0], current_bounds[1]]
            ax.scatter(start_bound, end_bound, color="blue", s=100, zorder=6, label="Bounds")
            
            title = graph_title.get() if graph_title else "Untitled"
            ax.set_title(title, fontsize=14, fontweight="bold")
            ax.set_xlabel("X")
            ax.set_ylabel("Value")
            ax.legend()
            
            fig.tight_layout()
            
            # Save figure temporarily
            img_path = filename.rsplit(".", 1)[0] + "_temp.png"
            fig.savefig(img_path, dpi=150, bbox_inches="tight")
            
            # Add image to Excel
            ws.add_image(XLImage(img_path), "D1")
            
            wb.save(filename)
            
            # Clean up temp image
            os.remove(img_path)
            
            messagebox.showinfo("Success", f"Graph and values saved to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel: {e}")

    def submit_edit(self):
        # Re-render the main graph after editing
        if self.current_bounds and self.current_values:
            self.plot_results(self.current_bounds, self.current_values)

    def render_plot(self, ax, canvas, bounds, plot_points, title):
        ax.cla()

        x_out = np.arange(len(plot_points))
        y_out = plot_points

        ax.plot(x_out, y_out, label="Interpolated (backend)")
        ax.scatter(x_out, y_out, color="red", zorder=5, label="Interpolated points")

        start_bound = [0, len(plot_points) - 1]
        end_bound = [bounds[0], bounds[1]]
        ax.scatter(start_bound, end_bound, color="blue", s=100, zorder=6, label="Bounds (click to edit)")

        ax.set_xlabel("X")
        ax.set_ylabel("Value")
        ax.set_title(title or "Value Interpolation Graph")

        ax.legend()
        ax.relim()
        ax.autoscale_view()

        canvas.draw()

    def plot_results(self, bounds, plot_points):
        # Call Value Generator
        if len(bounds) == 3 and plot_points == []:
            try:
                plot_points = ValueGenerator.interpolate_with_variation(*bounds)
            except Exception as e:
                messagebox.showerror("Backend error", f"Backend processing failed: {e}")
                return
        elif len(bounds) < 3:
            messagebox.showerror("Not enough points", "Please provide at least start, end, and steps values.")
            return

        self.current_bounds = bounds[:]
        self.current_values = plot_points[:]
        self.graph_title = tk.StringVar(value=self.graph_title.get())

        # Update main graph state
        self.main_graph_state["parent"] = self.root
        self.main_graph_state["ax"] = self.ax
        self.main_graph_state["canvas"] = self.canvas
        self.main_graph_state["bounds"] = self.current_bounds
        self.main_graph_state["plot_points"] = self.current_values
        self.main_graph_state["title"] = self.graph_title.get() or "Value Interpolation Graph"

        # Draw using shared renderer
        self.render_plot(
            self.ax,
            self.canvas,
            self.current_bounds,
            self.current_values,
            self.main_graph_state["title"]
        )

    def clear_all(self):
        # remove all entries and recreate start/end
        for lbl, ent in self.entries:
            lbl.destroy()
            ent.destroy()
        self.entries = []
        self._add_entry("Start")
        self._add_entry("End")
        self._add_entry("Steps")
        self._add_entry("Graph Title")
        self.current_values = []
        self.current_bounds = []
        self.graph_title = tk.StringVar()
        self._clear_plot()

    def close_child_window(self, window):
        try:
            if window in self.child_windows:
                self.child_windows.remove(window)
        except Exception:
            pass

        try:
            window.destroy()
        except Exception:
            pass 

    def on_app_close(self):
        """Cleanly close all app windows and terminate the process."""
        try:
            for window in self.child_windows[:]:
                try:
                    window.destroy()
                except Exception:
                    pass
            self.child_windows.clear()
        except Exception:
            pass

        try:
            plt.close("all")
        except Exception:
            pass

        try:
            self.root.quit()
        except Exception:
            pass

        try:
            self.root.destroy()
        except Exception:
            pass

        sys.exit(0)

def main():
    root = tk.Tk()
    InterpolationApp(root)
    root.geometry("700x700")
    root.mainloop()


if __name__ == "__main__":
    main()